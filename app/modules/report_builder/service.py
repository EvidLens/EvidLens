import os
from datetime import datetime
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from app.modules.market_engine.service import search_market
from app.modules.ai_insight.service import generate_ai_insight
from app.modules.knowledge_base.service import get_sector_benchmark

PRIMARY = (10, 31, 68) # #0A1F44 Dark Navy
SECONDARY = (20, 184, 166) #14B8A6 Teal
ACCENT = (245, 158, 11) #F59E0B Mustard

class EvidLensPDF(FPDF):
    def header(self):
        self.set_fill_color(*PRIMARY)
        self.rect(0, 0, 210, 30, 'F')
        self.set_text_color(255, 255, 255)
        self.set_font('Arial', 'B', 16)
        self.cell(0, 10, 'EvidLens', 0, 1, 'L')
        self.set_font('Arial', '', 10)
        self.cell(0, 6, 'Kenya’s Decision Intelligence Platform', 0, 1, 'L')
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_text_color(100, 100, 100)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Powered by EvidLens Kenya Sector Data | Page {self.page_no()}', 0, 0, 'C')

def generate_market_report_pdf(db: Session, query: str, sector: str, county: str, user_id: int):
    """Generate KRA-formatted PDF market report. Used for SME Starter + Pay-Per-Report"""
    market_data = search_market(db, query, sector, county)
    ai_insight = generate_ai_insight(db, query, sector, county)
    benchmark = get_sector_benchmark(sector)

    pdf = EvidLensPDF()
    pdf.add_page()
    pdf.set_text_color(0, 0, 0)

    # Title
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(0, 10, f'Market Feasibility Report: {query}', 0, 1)
    pdf.set_font('Arial', '', 9)
    pdf.cell(0, 6, f'Sector: {sector} | County: {county} | Generated: {datetime.now().strftime("%d %b %Y")}', 0, 1)
    pdf.ln(5)

    # Section 1: Executive Summary
    pdf.set_font('Arial', 'B', 12)
    pdf.set_text_color(*SECONDARY)
    pdf.cell(0, 8, '1. Executive Summary', 0, 1)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Arial', '', 10)
    pdf.multi_cell(0, 6, f"Recommendation: {ai_insight['recommendation']}\n\n{ai_insight['summary']}")
    pdf.ln(3)

    # Section 2: Market Metrics
    pdf.set_font('Arial', 'B', 12)
    pdf.set_text_color(*SECONDARY)
    pdf.cell(0, 8, '2. Market Metrics', 0, 1)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Arial', '', 10)
    pdf.cell(95, 6, f"Demand Level: {market_data['demand_level']}", 0, 0)
    pdf.cell(95, 6, f"Market Size: KES {market_data['market_size_kes']:,.0f}", 0, 1)
    pdf.cell(95, 6, f"Avg Price: KES {market_data['price_range']['avg_kes']:,.0f}", 0, 0)
    pdf.cell(95, 6, f"Competitors: {market_data['competitor_count']}", 0, 1)
    pdf.ln(3)

    # Section 3: Consumer Sentiment
    pdf.set_font('Arial', 'B', 12)
    pdf.set_text_color(*SECONDARY)
    pdf.cell(0, 8, '3. Consumer Sentiment', 0, 1)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Arial', '', 10)
    s = market_data['sentiment_summary']
    pdf.cell(0, 6, f"Positive: {s['positive']}% | Neutral: {s['neutral']}% | Negative: {s['negative']}%", 0, 1)
    pdf.multi_cell(0, 6, f"Common Complaints: {', '.join(s['complaints'])}")
    pdf.ln(3)

    # Section 4: Risk Analysis
    pdf.set_font('Arial', 'B', 12)
    pdf.set_text_color(*ACCENT)
    pdf.cell(0, 8, '4. Risk Analysis & Strategy', 0, 1)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Arial', '', 10)
    pdf.multi_cell(0, 6, f"Saturation Risk: {ai_insight['risk']}\nPricing Strategy: {ai_insight['pricing_strategy']}")
    pdf.ln(5)

    # KRA Compliance Footer
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(0, 8, 'KRA COMPLIANT REPORT - FOR BUSINESS PLANNING PURPOSES', 0, 1, 'C', 1)

    filename = f"evidlens_report_{user_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    filepath = f"/tmp/{filename}"
    pdf.output(filepath)
    return filepath

def generate_market_report_excel(db: Session, query: str, sector: str, county: str):
    """Generate Excel export for investors/banks. KRA formatted"""
    market_data = search_market(db, query, sector, county)
    ai_insight = generate_ai_insight(db, query, sector, county)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Market Report"

    # Header
    ws.merge_cells('A1:D1')
    ws['A1'] = f'EvidLens Market Report: {query}'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='0A1F44', end_color='0A1F44', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Data
    rows = [
        ["Metric", "Value", "County", "Sector"],
        ["Demand Level", market_data['demand_level'], county, sector],
        ["Market Size KES", market_data['market_size_kes'], county, sector],
        ["Avg Price KES", market_data['price_range']['avg_kes'], county, sector],
        ["Competitor Count", market_data['competitor_count'], county, sector],
        ["AI Recommendation", ai_insight['recommendation'], "", ""],
        ["Risk Level", ai_insight['risk'], "", ""]
    ]

    for row in rows:
        ws.append(row)

    # Styling
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='14B8A6', end_color='14B8A6', fill_type='solid')

    filename = f"evidlens_report_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    filepath = f"/tmp/{filename}"
    wb.save(filepath)
    return filepath
